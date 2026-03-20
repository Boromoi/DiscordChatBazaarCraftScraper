"""
Hypixel Skyblock – Bazaar Craft Scraper
========================================
Stuurt automatisch het /craft slash command in Discord,
scrapt alle pagina's, filtert op profit/volume,
en toont de resultaten in een GUI-venster.

Vereisten:
    pip install discord.py-self tabulate colorama openpyxl python-dotenv

Gebruik:
    1. Maak een .env bestand aan:
           DISCORD_USER_TOKEN=...
           CHANNEL_ID=...
           BOT_ID=...
    2. Stel de filters en opties in onder CONFIGURATIE.
    3. Draai het script — het stuurt automatisch het command.
"""

import re
import csv
import json
import asyncio
import tkinter as tk
import math
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, field

import discord
import requests
from colorama import Fore, Style, init

import os
from dotenv import load_dotenv

load_dotenv()
init(autoreset=True)

# ══════════════════════════════════════════════════════
#  CONFIGURATIE  –  pas dit aan
# ══════════════════════════════════════════════════════

USER_TOKEN  = os.getenv("DISCORD_USER_TOKEN", "")
CHANNEL_ID  = int(os.getenv("CHANNEL_ID", "0"))
BOT_ID      = int(os.getenv("BOT_ID", "0")) or None

# Slash command opties
BUY_METHOD  = "Buy Order"    # "Buy Order" of "Insta-Buy"
SELL_METHOD = "Sell Offer"   # "Sell Offer" of "Insta-Sell"

# Filters — zet op 0 om uit te schakelen
MIN_PROFIT  = 10000     # minimale profit in coins (bijv. 50000 = 50K)
MIN_VOLUME  = 100       # minimale volume in orders/week

DELAY        = 1.5      # seconden tussen pagina's
MAX_PAGES    = 50
PAGE_TIMEOUT = 30.0

# Gewichten combined score
WEIGHT_PROFIT = 0.4
WEIGHT_VOLUME = 0.4
WEIGHT_MARGIN = 0.2   # profit/input_cost % — beloont efficiency

# ══════════════════════════════════════════════════════
#  DATA MODEL
# ══════════════════════════════════════════════════════

@dataclass
class Material:
    amount: int
    name: str
    def __str__(self): return f"{self.amount}x {self.name}"


@dataclass
class Craft:
    rank:         int
    name:         str
    input_cost:   float
    output_value: float
    volume:       float
    profit:       float
    requires:     str   = ""
    materials:    list  = field(default_factory=list)
    score:        float = 0.0

    def profit_str(self) -> str: return _fmt(self.profit)
    def volume_str(self) -> str: return _fmt(self.volume)
    def cost_str(self)   -> str: return _fmt(self.input_cost)
    def output_str(self) -> str: return _fmt(self.output_value)
    def score_str(self)  -> str: return f"{self.score:.3f}"


# ══════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════

def _parse_coins(raw: str) -> float:
    raw = raw.strip().replace(",", "")
    for suffix, mult in (("B", 1_000_000_000), ("M", 1_000_000), ("K", 1_000)):
        if raw.endswith(suffix):
            return float(raw[:-1]) * mult
    return float(raw)


def _fmt(n: float) -> str:
    if n >= 1_000_000: return f"{n/1_000_000:.2f}M"
    if n >= 1_000:     return f"{n/1_000:.1f}K"
    return str(int(n))


def _timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d %H-%M-%S")


def _export_filename(sort_by: str, ext: str) -> Path:
    label = SORT_LABELS.get(sort_by, sort_by).split("(")[0].strip()
    label = re.sub(r"\s+", " ", label).strip()
    return Path(f"Bazaar Crafts - {label} - {_timestamp()}.{ext}")


# ══════════════════════════════════════════════════════
#  PARSER
# ══════════════════════════════════════════════════════

MATERIAL_RE = re.compile(r"(\d+)x\s+(.+)")


def parse_embed(message: discord.Message) -> list:
    if not message.embeds:
        return []

    crafts = []
    embed  = message.embeds[0]

    for field in embed.fields:
        name_match = re.match(r"(\d+)\.\s*(.+)", field.name or "")
        if not name_match:
            continue

        rank = int(name_match.group(1))
        name = re.sub(r"<:[^>]+>", "", name_match.group(2)).strip()
        val  = field.value or ""

        materials = []
        mat_block = re.search(r"```[a-z]*\n?([\s\S]+?)```", val)
        if mat_block:
            for line in mat_block.group(1).strip().splitlines():
                mat = MATERIAL_RE.match(line.strip())
                if mat:
                    materials.append(Material(int(mat.group(1)), mat.group(2).strip()))

        req_match = re.search(r"\*?\*?Requires:\*?\*?\s*(.+)", val)
        requires  = re.sub(r"<:[^>]+>", "", req_match.group(1)).strip() if req_match else ""

        cost_m  = re.search(r"\*\*Input cost:\*\*\s*([\d.,]+[KMBT]?)\s*coins", val)
        out_m   = re.search(r"\*\*Output value:\*\*\s*([\d.,]+[KMBT]?)\s*coins", val)
        vol_m   = re.search(r"\*\*Volume:\*\*\s*([\d.,]+[KMBT]?)\s*orders/week", val)
        prof_m  = re.search(r"\*\*Profit:\*\*\s*([\d.,]+[KMBT]?)\s*coins", val)

        if not (cost_m and out_m and vol_m and prof_m):
            continue

        profit = _parse_coins(prof_m.group(1))
        volume = _parse_coins(vol_m.group(1))

        # Filters toepassen
        if profit < MIN_PROFIT or volume < MIN_VOLUME:
            continue

        crafts.append(Craft(
            rank         = rank,
            name         = name,
            input_cost   = _parse_coins(cost_m.group(1)),
            output_value = _parse_coins(out_m.group(1)),
            volume       = volume,
            profit       = profit,
            requires     = requires,
            materials    = materials,
        ))

    return crafts


def get_page_info_embed(message: discord.Message) -> tuple:
    if not message.embeds:
        return 1, 1
    embed = message.embeds[0]
    footer_text = embed.footer.text if embed.footer else ""
    m = re.search(r"Page (\d+)[/•\s]+(\d+)", footer_text or "")
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.search(r"Page (\d+)[/•\s]+(\d+)", embed.description or "")
    if m:
        return int(m.group(1)), int(m.group(2))
    return 1, 1


def get_button(message: discord.Message, label_keywords: list):
    for row in message.components:
        for comp in row.children:
            if isinstance(comp, discord.Button):
                label = (comp.label or "").lower()
                if any(kw in label for kw in label_keywords):
                    return comp
    return None


# ══════════════════════════════════════════════════════
#  COMBINED SCORE
# ══════════════════════════════════════════════════════

def compute_scores(crafts: list) -> None:
    if not crafts:
        return

    def log_norm(values: list) -> list:
        log_vals = [math.log10(max(v, 1)) for v in values]
        mn, mx = min(log_vals), max(log_vals)
        r = mx - mn or 1
        return [(v - mn) / r for v in log_vals]

    profits = [c.profit for c in crafts]
    volumes = [c.volume for c in crafts]
    margins = [c.profit / max(c.input_cost, 1) * 100 for c in crafts]

    lp = log_norm(profits)
    lv = log_norm(volumes)
    lm = log_norm(margins)

    for i, c in enumerate(crafts):
        c.score = WEIGHT_PROFIT * lp[i] + WEIGHT_VOLUME * lv[i] + WEIGHT_MARGIN * lm[i]

# ══════════════════════════════════════════════════════
#  SORTERING
# ══════════════════════════════════════════════════════

SORT_KEYS = {
    "profit":   lambda c: c.profit,
    "volume":   lambda c: c.volume,
    "combined": lambda c: c.score,
    "cost":     lambda c: c.input_cost,
    "output":   lambda c: c.output_value,
    "rank":     lambda c: c.rank,
}
SORT_LABELS = {
    "profit":   "Profit",
    "volume":   "Volume",
    "combined": f"Beste combinatie (profit {int(WEIGHT_PROFIT*100)}% + volume {int(WEIGHT_VOLUME*100)}%)",
    "cost":     "Input cost",
    "output":   "Output value",
    "rank":     "Rank",
}


def sort_crafts(crafts: list, sort_by: str) -> list:
    return sorted(crafts, key=SORT_KEYS.get(sort_by, SORT_KEYS["profit"]),
                  reverse=(sort_by != "rank"))


# ══════════════════════════════════════════════════════
#  SLASH COMMAND VERSTUREN
# ══════════════════════════════════════════════════════

async def send_craft_command(client: discord.Client, channel: discord.TextChannel) -> bool:
    """
    Stuurt het /craft slash command via de Discord interaction API.
    Haalt eerst de command info op, dan POST naar /interactions.
    """
    print(Fore.CYAN + f"  /craft command ophalen...")

    headers = {
        "Authorization": USER_TOKEN,
        "Content-Type":  "application/json",
        "User-Agent":    "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
    }

    # Zoek het /craft command via de search endpoint
    search_url = (
        f"https://discord.com/api/v9/channels/{CHANNEL_ID}"
        f"/application-commands/search?type=1&query=craft&limit=10&include_applications=true"
    )
    resp = requests.get(search_url, headers=headers)
    if resp.status_code != 200:
        print(Fore.RED + f"  Command ophalen mislukt: {resp.status_code}")
        return False

    data         = resp.json()
    commands     = data.get("application_commands", [])
    applications = {a["id"]: a for a in data.get("applications", [])}

    craft_cmd = next((c for c in commands if c.get("name") == "craft"), None)
    if not craft_cmd:
        print(Fore.RED + "  /craft command niet gevonden in dit kanaal.")
        return False

    cmd_id     = craft_cmd["id"]
    app_id     = craft_cmd["application_id"]
    guild_id   = str(channel.guild.id)

    # Bouw de opties
    # Zoek de optie IDs/choices op uit de command definitie
    options_map = {opt["name"]: opt for opt in craft_cmd.get("options", [])}

    def find_choice_value(option_name: str, label: str) -> str:
        opt = options_map.get(option_name, {})
        for choice in opt.get("choices", []):
            if choice["name"].lower() == label.lower():
                return choice["value"]
        return label  # fallback: gebruik label als value

    buy_value  = find_choice_value("buymethod",  BUY_METHOD)
    sell_value = find_choice_value("sellmethod", SELL_METHOD)

    payload = {
        "type":           2,
        "application_id": app_id,
        "guild_id":       guild_id,
        "channel_id":     str(CHANNEL_ID),
        "session_id":     "scraper_session",
        "data": {
            "version":            craft_cmd.get("version", "1"),
            "id":                 cmd_id,
            "name":               "craft",
            "type":               1,
            "options": [
                {"type": 3, "name": "buymethod",  "value": buy_value},
                {"type": 3, "name": "sellmethod", "value": sell_value},
            ],
        },
        "nonce": str(int(datetime.now().timestamp() * 1000)),
    }

    resp = requests.post(
        "https://discord.com/api/v9/interactions",
        headers=headers, json=payload,
    )

    if resp.status_code in (200, 204):
        print(Fore.GREEN + f"  ✓ /craft {BUY_METHOD} / {SELL_METHOD} verstuurd!")
        return True
    else:
        print(Fore.RED + f"  Command versturen mislukt: {resp.status_code} — {resp.text[:200]}")
        return False


# ══════════════════════════════════════════════════════
#  DISCORD CLIENT
# ══════════════════════════════════════════════════════

class CraftScraper(discord.Client):

    def __init__(self):
        super().__init__()
        self.target_channel_id = CHANNEL_ID
        self.bot_id            = BOT_ID
        self.all_crafts: list  = []
        self.seen_ranks: set   = set()
        self._scraping         = False
        self._tracked_message  = None
        self._page_event       = asyncio.Event()
        self._command_sent     = False

    def _get_content(self, message: discord.Message) -> str:
        if message.embeds:
            embed = message.embeds[0]
            parts = []
            if embed.title:       parts.append(embed.title)
            if embed.description: parts.append(embed.description)
            for f in embed.fields:
                if f.name:  parts.append(f.name)
                if f.value: parts.append(f.value)
            return "\n".join(parts)
        return message.content or ""

    def _is_craft_message(self, message: discord.Message) -> bool:
        if not message.embeds:
            return False
        title = re.sub(r"<:[^>]+>", "", message.embeds[0].title or "").strip()
        return "Profitable Bazaar Crafts" in title

    async def on_ready(self):
        print(Fore.GREEN + f"  ✓ Ingelogd als {self.user}")
        channel = self.get_channel(self.target_channel_id)
        if not channel:
            print(Fore.RED + "  Kanaal niet gevonden!")
            await self.close()
            return

        await asyncio.sleep(1)
        success = await send_craft_command(self, channel)
        if not success:
            print(Fore.YELLOW + "  Command versturen mislukt — typ het handmatig in Discord.")
        self._command_sent = True

    async def on_message(self, message: discord.Message):
        if message.channel.id != self.target_channel_id:
            return
        if not message.author.bot:
            return
        if self.bot_id and message.author.id != self.bot_id:
            return
        if self._scraping:
            return
        if not self._is_craft_message(message):
            return

        self._scraping        = True
        self._tracked_message = message
        print(Fore.CYAN + f"\n  Craft data ontvangen van {message.author}!")
        asyncio.create_task(self._scrape(message))

    async def on_message_edit(self, before: discord.Message, after: discord.Message):
        if after.channel.id != self.target_channel_id:
            return
        if not after.author.bot:
            return
        if self.bot_id and after.author.id != self.bot_id:
            return
        if not self._is_craft_message(after):
            return

        if self._tracked_message and after.id == self._tracked_message.id:
            self._tracked_message = after
            self._page_event.set()
            if self._scraping:
                return

        if self._scraping:
            return

        print(Fore.CYAN + f"\n  Craft data ontvangen via edit van {after.author}!")
        self._scraping        = True
        self._tracked_message = after
        asyncio.create_task(self._scrape(after))

    async def _scrape(self, message: discord.Message):
        print(Fore.CYAN + "  Scrapen gestart...\n")

        for _ in range(MAX_PAGES):
            cur, _ = get_page_info_embed(message)
            if cur <= 1:
                break
            prev_btn = get_button(message, ["prev", "←", "◀", "<"])
            if not prev_btn:
                break
            self._page_event.clear()
            await prev_btn.click()
            message = await self._wait_for_update()
            if not message:
                break

        for _ in range(MAX_PAGES):
            cur, total = get_page_info_embed(message)
            print(Fore.GREEN + f"  Pagina {cur}/{total}...", end=" ", flush=True)

            new = 0
            for craft in parse_embed(message):
                if craft.rank not in self.seen_ranks:
                    self.all_crafts.append(craft)
                    self.seen_ranks.add(craft.rank)
                    new += 1
            print(f"({new} crafts)")

            if cur >= total:
                print(Fore.CYAN + "  Laatste pagina bereikt.")
                break

            next_btn = get_button(message, ["next", "→", "▶", ">"])
            if not next_btn:
                print(Fore.RED + "  Geen 'next' knop gevonden, stoppen.")
                break

            self._page_event.clear()
            await next_btn.click()
            message = await self._wait_for_update()
            if not message:
                print(Fore.RED + "  Timeout — stoppen.")
                break

        compute_scores(self.all_crafts)
        count = len(self.all_crafts)
        print(Fore.GREEN + Style.BRIGHT + f"\n  ✓ {count} crafts gescraped & gescoord!")
        if MIN_PROFIT > 0 or MIN_VOLUME > 0:
            print(Fore.YELLOW + f"  Filter: profit ≥ {_fmt(MIN_PROFIT)}  |  volume ≥ {_fmt(MIN_VOLUME)}")
        await self.close()

    async def _wait_for_update(self):
        await asyncio.sleep(DELAY)
        try:
            await asyncio.wait_for(self._page_event.wait(), timeout=PAGE_TIMEOUT)
        except asyncio.TimeoutError:
            return None
        return self._tracked_message


# ══════════════════════════════════════════════════════
#  EXPORTS
# ══════════════════════════════════════════════════════

def export_csv(crafts: list, sort_by: str) -> Path:
    path = _export_filename(sort_by, "csv")
    with path.open("w", newline="", encoding="utf-8") as f:
        fieldnames = ["rank", "name", "requires", "input_cost",
                      "output_value", "volume", "profit", "score", "materials"]
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for c in sort_crafts(crafts, sort_by):
            w.writerow({
                "rank": c.rank, "name": c.name, "requires": c.requires,
                "input_cost": c.input_cost, "output_value": c.output_value,
                "volume": c.volume, "profit": c.profit,
                "score": round(c.score, 4),
                "materials": " | ".join(str(m) for m in c.materials),
            })
    return path


def export_json(crafts: list, sort_by: str) -> Path:
    path = _export_filename(sort_by, "json")
    data = [{
        "rank": c.rank, "name": c.name, "requires": c.requires,
        "input_cost": c.input_cost, "output_value": c.output_value,
        "volume": c.volume, "profit": c.profit, "score": round(c.score, 4),
        "materials": [{"amount": m.amount, "name": m.name} for m in c.materials],
    } for c in sort_crafts(crafts, sort_by)]
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    return path


def export_excel(crafts: list, sort_by: str) -> Path:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    path = _export_filename(sort_by, "xlsx")
    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = "Bazaar Crafts"

    hdr_fill   = PatternFill("solid", fgColor="1E3A5F")
    hdr_font   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    even_fill  = PatternFill("solid", fgColor="E8F5E9")
    odd_fill   = PatternFill("solid", fgColor="F5F5F5")
    score_fill = PatternFill("solid", fgColor="FFF9C4")
    b          = Side(style="thin", color="CCCCCC")
    border     = Border(left=b, right=b, top=b, bottom=b)
    right      = Alignment(horizontal="right", vertical="center")
    center     = Alignment(horizontal="center", vertical="center")

    col_headers = ["#", "Craft naam", "Requires", "Input cost (coins)",
                   "Output value (coins)", "Volume/week",
                   "Score (combined)", "Profit (coins)", "Materials"]
    col_widths  = [5, 32, 20, 20, 22, 14, 16, 20, 52]

    for col, (h, w) in enumerate(zip(col_headers, col_widths), 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    for ri, c in enumerate(sort_crafts(crafts, sort_by), 2):
        base = even_fill if ri % 2 == 0 else odd_fill
        row  = [c.rank, c.name + (f" [{c.requires}]" if c.requires else ""),
                c.requires, c.input_cost, c.output_value, c.volume,
                round(c.score, 4), c.profit,
                " | ".join(str(m) for m in c.materials)]
        for ci, val in enumerate(row, 1):
            cell        = ws.cell(row=ri, column=ci, value=val)
            cell.fill   = score_fill if ci == 7 else base
            cell.border = border
            if ci in (1, 4, 5, 6, 7, 8):
                cell.alignment = right
            if ci in (4, 5, 8):
                cell.number_format = "#,##0"
            if ci == 7:
                cell.number_format = "0.000"

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = f"A1:I{ws.max_row}"
    tr = ws.max_row + 1
    ws.cell(row=tr, column=1, value="TOTAAL").font = Font(bold=True)
    ws.cell(row=tr, column=8,
            value=f"=SUM(H2:H{tr-1})").number_format = "#,##0"
    wb.save(path)
    return path


def export_markdown(crafts: list, sort_by: str) -> Path:
    path  = _export_filename(sort_by, "md")
    label = SORT_LABELS[sort_by]
    lines = [
        "# Hypixel Bazaar Crafts",
        f"_Geëxporteerd op {datetime.now().strftime('%d-%m-%Y %H:%M')} · {label}_\n",
        "| # | Craft | Requires | Score | Cost | Output | Volume/wk | Profit | Materials |",
        "|---|-------|----------|-------|------|--------|-----------|--------|-----------|",
    ]
    for c in sort_crafts(crafts, sort_by):
        mats = " \\| ".join(str(m) for m in c.materials)
        lines.append(
            f"| {c.rank} | {c.name} | {c.requires} | {c.score_str()} | "
            f"{c.cost_str()} | {c.output_str()} | {c.volume_str()} | {c.profit_str()} | {mats} |"
        )
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


# ══════════════════════════════════════════════════════
#  GUI
# ══════════════════════════════════════════════════════

COLUMNS = [
    ("rank",     "#",          60,  "center"),
    ("name",     "Craft",      260, "w"),
    ("cost",     "Cost",       90,  "e"),
    ("output",   "Output",     90,  "e"),
    ("volume",   "Volume/wk",  100, "e"),
    ("score",    "Score",      70,  "center"),
    ("profit",   "Profit",     100, "e"),
    ("mats",     "Materials",  300, "w"),
]

SORT_COL_MAP = {
    "rank":   "rank",
    "name":   "rank",
    "cost":   "cost",
    "output": "output",
    "volume": "volume",
    "score":  "combined",
    "profit": "profit",
    "mats":   "rank",
}


class CraftGUI:
    def __init__(self, crafts: list):
        self.crafts      = crafts
        self.sort_by     = "profit"
        self.sort_rev    = True

        self.root = tk.Tk()
        self.filter_profit = tk.StringVar(value=str(int(MIN_PROFIT)))
        self.filter_volume = tk.StringVar(value=str(int(MIN_VOLUME)))
        self.root.title("Hypixel Bazaar Craft Scraper")
        self.root.geometry("1400x750")
        self.root.configure(bg="#1a1a2e")

        self._build_ui()
        self._load_table(self.crafts)

    def _build_ui(self):
        # ── Header ──────────────────────────────────
        hdr = tk.Frame(self.root, bg="#16213e", pady=8)
        hdr.pack(fill="x")
        tk.Label(hdr, text="🪙  Hypixel Bazaar Craft Scraper",
                 font=("Segoe UI", 16, "bold"), fg="#e2b96f", bg="#16213e").pack(side="left", padx=16)
        self.status_label = tk.Label(hdr, text=f"{len(self.crafts)} crafts",
                                     font=("Segoe UI", 11), fg="#aaaaaa", bg="#16213e")
        self.status_label.pack(side="right", padx=16)

        # ── Filter bar ──────────────────────────────
        fbar = tk.Frame(self.root, bg="#0f3460", pady=6)
        fbar.pack(fill="x")

        tk.Label(fbar, text="Min Profit:", fg="white", bg="#0f3460",
                 font=("Segoe UI", 10)).pack(side="left", padx=(12, 4))
        tk.Entry(fbar, textvariable=self.filter_profit, width=10,
                 font=("Segoe UI", 10)).pack(side="left", padx=(0, 12))

        tk.Label(fbar, text="Min Volume/wk:", fg="white", bg="#0f3460",
                 font=("Segoe UI", 10)).pack(side="left", padx=(0, 4))
        tk.Entry(fbar, textvariable=self.filter_volume, width=10,
                 font=("Segoe UI", 10)).pack(side="left", padx=(0, 12))

        tk.Button(fbar, text="Toepassen", command=self._apply_filters,
                  bg="#e2b96f", fg="#1a1a2e", font=("Segoe UI", 10, "bold"),
                  relief="flat", padx=10).pack(side="left")

        # ── Export buttons ──────────────────────────
        exp_frame = tk.Frame(fbar, bg="#0f3460")
        exp_frame.pack(side="right", padx=12)
        for label, fn in [("CSV", self._export_csv), ("JSON", self._export_json),
                           ("Excel", self._export_excel), ("Markdown", self._export_md)]:
            tk.Button(exp_frame, text=label, command=fn,
                      bg="#533483", fg="white", font=("Segoe UI", 9, "bold"),
                      relief="flat", padx=8, pady=2).pack(side="left", padx=3)

        # ── Treeview ────────────────────────────────
        tree_frame = tk.Frame(self.root, bg="#1a1a2e")
        tree_frame.pack(fill="both", expand=True, padx=8, pady=8)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview",
                         background="#1a1a2e", foreground="white",
                         fieldbackground="#1a1a2e", rowheight=24,
                         font=("Segoe UI", 10))
        style.configure("Treeview.Heading",
                         background="#16213e", foreground="#e2b96f",
                         font=("Segoe UI", 10, "bold"), relief="flat")
        style.map("Treeview", background=[("selected", "#0f3460")])

        self.tree = ttk.Treeview(tree_frame, columns=[c[0] for c in COLUMNS],
                                  show="headings", selectmode="browse")

        for col_id, col_label, col_width, col_anchor in COLUMNS:
            self.tree.heading(col_id, text=col_label,
                              command=lambda c=col_id: self._sort_column(c))
            self.tree.column(col_id, width=col_width, anchor=col_anchor, minwidth=40)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # Alternerende rijkleuren
        self.tree.tag_configure("odd",  background="#16213e")
        self.tree.tag_configure("even", background="#1a1a2e")
        self.tree.tag_configure("top",  background="#1b3a1b")  # groen voor top 10

        # Tooltip bij dubbelklik
        self.tree.bind("<Double-1>", self._show_detail)

    def _load_table(self, crafts: list):
        self.tree.delete(*self.tree.get_children())
        sorted_crafts = sort_crafts(crafts, self.sort_by)
        for i, c in enumerate(sorted_crafts):
            mats = " | ".join(str(m) for m in c.materials)
            tag  = "top" if i < 10 else ("even" if i % 2 == 0 else "odd")
            self.tree.insert("", "end", iid=str(c.rank), tags=(tag,), values=(
                c.rank, c.name + (f" [{c.requires}]" if c.requires else ""),
                c.cost_str(), c.output_str(), c.volume_str(),
                c.score_str(), c.profit_str(), mats,
            ))
        self.status_label.config(text=f"{len(crafts)} crafts  |  sort: {SORT_LABELS[self.sort_by]}")

    def _sort_column(self, col_id: str):
        new_sort = SORT_COL_MAP.get(col_id, "profit")
        if new_sort == self.sort_by:
            self.sort_rev = not self.sort_rev
        else:
            self.sort_by  = new_sort
            self.sort_rev = True
        self._load_table(self._filtered_crafts())

    def _filtered_crafts(self) -> list:
        try:
            min_p = float(self.filter_profit.get() or 0)
            min_v = float(self.filter_volume.get() or 0)
        except ValueError:
            min_p = min_v = 0
        return [c for c in self.crafts if c.profit >= min_p and c.volume >= min_v]

    def _apply_filters(self):
        self._load_table(self._filtered_crafts())

    def _show_detail(self, event):
        item = self.tree.focus()
        if not item:
            return
        rank = int(item)
        craft = next((c for c in self.crafts if c.rank == rank), None)
        if not craft:
            return
        mats = "\n".join(f"  • {m}" for m in craft.materials)
        info = (
            f"Craft #{craft.rank}: {craft.name}\n"
            f"{'─'*40}\n"
            f"Requires:  {craft.requires or '—'}\n"
            f"Cost:      {craft.cost_str()}\n"
            f"Output:    {craft.output_str()}\n"
            f"Volume:    {craft.volume_str()} orders/week\n"
            f"Profit:    {craft.profit_str()}\n"
            f"Score:     {craft.score_str()}\n"
            f"{'─'*40}\n"
            f"Materials:\n{mats}"
        )
        messagebox.showinfo(craft.name, info)

    # Export helpers
    def _do_export(self, fn, ext):
        path = fn(self._filtered_crafts(), self.sort_by)
        if path:
            messagebox.showinfo("Export geslaagd", f"Opgeslagen als:\n{path.resolve()}")

    def _export_csv(self):   self._do_export(export_csv,      "csv")
    def _export_json(self):  self._do_export(export_json,     "json")
    def _export_excel(self): self._do_export(export_excel,    "xlsx")
    def _export_md(self):    self._do_export(export_markdown, "md")

    def run(self):
        self.root.mainloop()


# ══════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════

if __name__ == "__main__":
    print()
    print(Fore.CYAN + Style.BRIGHT + "  ╔════════════════════════════════════════╗")
    print(Fore.CYAN + Style.BRIGHT + "  ║   Hypixel Bazaar Craft Scraper  🪙    ║")
    print(Fore.CYAN + Style.BRIGHT + "  ╚════════════════════════════════════════╝")
    print()
    print(Fore.YELLOW + f"  Buy: {BUY_METHOD}  |  Sell: {SELL_METHOD}")
    if MIN_PROFIT > 0 or MIN_VOLUME > 0:
        print(Fore.YELLOW + f"  Filter: profit ≥ {_fmt(MIN_PROFIT)}  |  volume ≥ {_fmt(MIN_VOLUME)}")
    print()

    if not USER_TOKEN:
        print(Fore.RED + "  Geen token gevonden!")
        exit(1)

    client = CraftScraper()
    client.run(USER_TOKEN)

    if not client.all_crafts:
        print(Fore.RED + "\n  Geen crafts gevonden.")
    else:
        print(Fore.CYAN + "\n  GUI openen...")
        gui = CraftGUI(client.all_crafts)
        gui.run()