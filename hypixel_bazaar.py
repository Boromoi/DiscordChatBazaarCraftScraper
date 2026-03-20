"""
Hypixel SkyBlock – Bazaar Analysis Tool
=========================================
Analyseert via de officiële Hypixel API:
  1. Craft Flips    – koop materials (buy order) → craft → verkoop (sell offer)
  2. Bazaar Flips   – koop (buy order) → verkoop (sell offer), spread trading
  3. AH Craft Flips – koop materials (bazaar) → craft → verkoop op Auction House

Vereisten:
    pip install requests openpyxl python-dotenv

Gebruik:
    1. Maak .env aan:  HYPIXEL_API_KEY=jouw_key_hier
    2. API key:  join Hypixel → /api new in-game
    3. python hypixel_bazaar.py
"""

import re
import csv
import json
import math
import time
import zipfile
import io
import threading
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
from pathlib import Path
from dataclasses import dataclass, field
from collections import defaultdict

import requests
import os
from dotenv import load_dotenv

load_dotenv()

# ══════════════════════════════════════════════════════
#  CONFIGURATIE
# ══════════════════════════════════════════════════════

HYPIXEL_API_KEY = os.getenv("HYPIXEL_API_KEY", "")

# Belasting
BAZAAR_SELL_TAX = 0.0125   # 1.25% op sell offers
AH_TAX          = 0.02     # ~2% totale AH tax

# Craft flip filters
MIN_CRAFT_PROFIT = 10_000
MIN_CRAFT_VOLUME = 100

# Bazaar flip filters
MIN_FLIP_PROFIT  = 50_000
MIN_FLIP_MARGIN  = 1.0     # minimale margin % (spread / buy_price)
MIN_FLIP_VOLUME  = 10_000  # orders/week

# AH craft filters
MIN_AH_PROFIT    = 100_000
MIN_AH_SALES     = 5       # minimale verkopen in recente data

# Score gewichten (moeten samen 1.0 zijn per modus)
WEIGHT_PROFIT = 0.4
WEIGHT_VOLUME = 0.4
WEIGHT_MARGIN = 0.2

# Cache instellingen
RECIPE_CACHE_FILE  = Path("recipes_cache.json")
RECIPE_CACHE_HOURS = 24
AH_CACHE_FILE      = Path("ah_cache.json")
AH_CACHE_MINUTES   = 10

# NEU repo
NEU_ZIP_URL = "https://github.com/NotEnoughUpdates/NotEnoughUpdates-REPO/archive/refs/heads/master.zip"

# ══════════════════════════════════════════════════════
#  DATA CLASSES
# ══════════════════════════════════════════════════════

@dataclass
class BazaarItem:
    item_id:          str
    display_name:     str
    buy_price:        float   # laagste sell offer (insta-buy prijs)
    sell_price:       float   # hoogste buy order (insta-sell prijs)
    buy_volume:       int     # orders/week kopen
    sell_volume:      int     # orders/week verkopen

    def spread(self) -> float:
        return self.buy_price - self.sell_price

    def margin_pct(self) -> float:
        if self.buy_price == 0:
            return 0
        return self.spread() / self.buy_price * 100


@dataclass
class CraftFlip:
    """Koop materials op bazaar → craft → verkoop op bazaar."""
    name:         str
    item_id:      str
    input_cost:   float         # kosten voor materials (buy order prijs)
    output_value: float         # opbrengst verkoop (sell offer - tax)
    profit:       float
    margin_pct:   float
    volume:       float         # orders/week van het gecraftte item
    ingredients:  list = field(default_factory=list)  # [(naam, hoeveelheid, prijs)]
    score:        float = 0.0

    def profit_str(self)  -> str: return _fmt(self.profit)
    def cost_str(self)    -> str: return _fmt(self.input_cost)
    def output_str(self)  -> str: return _fmt(self.output_value)
    def volume_str(self)  -> str: return _fmt(self.volume)
    def margin_str(self)  -> str: return f"{self.margin_pct:.1f}%"
    def score_str(self)   -> str: return f"{self.score:.3f}"


@dataclass
class BazaarFlip:
    """Koop via buy order → verkoop via sell offer (spread trading)."""
    item_id:         str
    name:            str
    buy_order:       float    # prijs buy order plaatsen
    sell_offer:      float    # prijs sell offer plaatsen
    profit:          float    # profit per item na tax
    margin_pct:      float
    volume:          float    # min(buy_volume, sell_volume) orders/week
    buy_volume:      float = 0.0   # coins gekocht via buy orders/week
    sell_volume:     float = 0.0   # coins verkocht via sell offers/week
    score:           float = 0.0

    def profit_str(self)      -> str: return _fmt(self.profit)
    def buy_str(self)         -> str: return _fmt(self.buy_order)
    def sell_str(self)        -> str: return _fmt(self.sell_offer)
    def volume_str(self)      -> str: return _fmt(self.volume)
    def buy_vol_str(self)     -> str: return _fmt(self.buy_volume)
    def sell_vol_str(self)    -> str: return _fmt(self.sell_volume)
    def margin_str(self)      -> str: return f"{self.margin_pct:.1f}%"
    def score_str(self)       -> str: return f"{self.score:.3f}"


@dataclass
class AHCraftFlip:
    """Koop materials op bazaar → craft → verkoop op AH."""
    name:          str
    item_id:       str
    material_cost: float        # kosten materials (insta-buy)
    ah_price:      float        # mediaan AH verkoopprijs
    profit:        float        # ah_price * (1-tax) - material_cost
    margin_pct:    float
    ah_sales:      int          # aantal verkopen in recente data
    ingredients:   list = field(default_factory=list)
    score:         float = 0.0

    def profit_str(self)  -> str: return _fmt(self.profit)
    def cost_str(self)    -> str: return _fmt(self.material_cost)
    def ah_str(self)      -> str: return _fmt(self.ah_price)
    def margin_str(self)  -> str: return f"{self.margin_pct:.1f}%"
    def score_str(self)   -> str: return f"{self.score:.3f}"


# ══════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════

INGREDIENT_ID_MAP = {
    # ── Vis / inkt varianten (NEU damage suffix) ──────────────────────────
    "RAW_FISH-1":               "ENCHANTED_RAW_SALMON",
    "RAW_FISH-2":               "RAW_FISH",
    "RAW_FISH-3":               "ENCHANTED_PUFFERFISH",
    "INK_SACK-3":               "INK_SACK",
    "INK_SACK-4":               "INK_SACK",
    "INK_SACK-15":              "INK_SACK",          # Bonemeal variant

    # ── Hout / log varianten (NEU gebruikt damage suffix) ─────────────────
    "WOOD":                     "LOG",               # Oak Log
    "WOOD-1":                   "LOG",               # Oak (variant)
    "WOOD-2":                   "SPRUCE_LOG",
    "WOOD-3":                   "BIRCH_LOG",
    "WOOD-4":                   "JUNGLE_LOG",
    "WOOD-5":                   "ACACIA_LOG",
    "WOOD-6":                   "DARK_OAK_LOG",
    "LOG-1":                    "SPRUCE_LOG",
    "LOG-2":                    "DARK_OAK_LOG",
    "LOG-3":                    "JUNGLE_LOG",
    "LOG_2":                    "ACACIA_LOG",
    "LOG_2-1":                  "DARK_OAK_LOG",

    # ── Naam verschillen NEU vs bazaar ────────────────────────────────────
    "SAND-1":                   "SAND",              # Red Sand → gewone zand als fallback
    "HARD_CLAY":                "CLAY_BALL",         # Clay Ball in bazaar
    "NETHER_STALK":             "NETHER_STALK",      # Nether Wart Misschien nog hernoemen
    "ENCHANTED_NETHER_STALK":   "ENCHANTED_NETHER_STALK",
    "WATER_LILY":               "WATER_LILY",        # Lily Pad Misschien nog hernoemen
    "ENCHANTED_WATER_LILY":     "ENCHANTED_WATER_LILY",
    "ENCHANTED_COCOA":          "ENCHANTED_COCOA",   # Enchanted Cocoa Beans
    "ENCHANTED_MELON_BLOCK":    "ENCHANTED_MELON",   # Enkelvoud in bazaar
    "ENCHANTED_ENDSTONE":       "ENCHANTED_END_STONE",
    "ENDER_STONE":              "END_STONE",
    "DOUBLE_PLANT":             "DOUBLE_PLANT",      # Sunflower / tall grass
    "POTATO_ITEM":              "POTATO_ITEM",
    "CARROT_ITEM":              "CARROT_ITEM",

    # ── Rift-only / unobtainables → prijs 0, recept blijft geldig ─────────
    # (worden afgehandeld via ZERO_COST_ITEMS hieronder)
}

# Items die niet op bazaar/AH/NPC te kopen zijn maar toch in recepten staan.
# We geven ze prijs 0 zodat het recept niet wordt weggegooid maar de kosten
# ook niet worden opgeblazen. Zet ze gewoon als bekende kosten van 0.
ZERO_COST_ITEMS = {
    "LIVING_METAL",             # Rift only
    "YOUNGITE",                 # Rift only
    "COVEN_SEAL",               # Rift only (Vampire Slayer drop)
    "BLOODBADGE",               # Rift only (Vampire Slayer drop)
    #"ZOMBIE_HEART",             # AH only sell item
    "LEECH_SUPREME_FRAGMENT",   # Rift only
    "METAL_HEART",              # Rift only (Living Metal Heart)
    #"GIANT_FRAGMENT_BIGFOOT",   # Bigfoot's Bola fragment
    #"GIANT_FRAGMENT_BOULDER",   # Jolly Pink Rock fragment

    "WILTED_BERBERIS",          # Rift only
    "TIMITE",                   # Rift only
    "FROSTY_CRUX",              # Rift only
    #"COLOSSAL_EXP_BOTTLE_UPGRADE",   # Upgrade item AH
    #"ULTIMATE_CARROT_CANDY_UPGRADE", # Upgrade item AH 
    "OBSOLITE",                 # NEU typo / onbekend item
}

# Vaste NPC prijzen voor items die niet in de bazaar zitten
# Bron: Hypixel SkyBlock NPC shops
NPC_PRICES = {
    "GLASS_BOTTLE":             3,
    #"HEAT_CORE":                200_000,    # Forging item, koopt via NPC (~200K)
    #"MINION_STORAGE_EXPANDER":  500_000,    # Dungeon shop item
    "BOWL":                     4,
    "STICK":                    4,
    "BLAZE_POWDER":             80,
    "PAPER":                    6,
    "GOLD_BLOCK":               204,   # 9x gold ingot à ~20 coins
    "GOLD_INGOT":               24,
    "IRON_INGOT":               6,
    "WOOD_SWORD":               8,
    "COAL":                     3,
    "FEATHER":                  4,
    "FLINT":                    2,
    "STRING":                   4,
    "ARROW":                    5,
    "BONE":                     6,
    "ROTTEN_FLESH":             2,
    "SPIDER_EYE":               12,
    "GUNPOWDER":                7,
    "ENDER_PEARL":              12,
    "BLAZE_ROD":                80,
    "MAGMA_CREAM":              16,
    "GHAST_TEAR":               400,
    "NETHER_BRICK":             8,
    "QUARTZ":                   4,
    "PRISMARINE_SHARD":         6,
    "PRISMARINE_CRYSTALS":      8,
    "OAK_LOG":                  2,
    "SAND":                     2,
    "GRAVEL":                   2,
    "CLAY_BALL":                3,
    "CACTUS":                   4,
    "SUGAR_CANE":               2,
    "POTATO_ITEM":              2,
    "CARROT_ITEM":              2,
    "WHEAT":                    2,
    "SEEDS":                    1,
    "MELON":                    2,
    "PUMPKIN":                  4,
    "RED_MUSHROOM":             4,
    "BROWN_MUSHROOM":           4,
    "COBBLESTONE":              2,
    "NETHERRACK":               1,
    "OBSIDIAN":                 4,
    "ICE":                      3,
    "SNOW_BALL":                1,
    "WATER_LILY":               4, #misschien nog naam veranderen
    "SPONGE":                   60,
    "RAW_FISH":                 4,
    "RAW_SALMON":               6,
    "PUFFERFISH":               6,
    "INK_SACK":                 8,
    "EGG":                      3,
    "RABBIT":                   6,
    "RABBIT_HIDE":              3,
    "RAW_BEEF":                 4,
    "RAW_CHICKEN":              4,
    "PORK":                     4,
    "MUTTON":                   4,
}


def _normalize_ingredient_id(ing_id: str) -> str:
    """Zet NEU ingredient ID om naar Hypixel bazaar ID."""
    return INGREDIENT_ID_MAP.get(ing_id, ing_id)


def _get_ingredient_price(ing_id: str, bazaar: dict,
                          ah_prices: dict = None) -> tuple[float, str]:
    """
    Geeft (prijs, display_naam) van een ingredient.
    Volgorde: bazaar → NPC → AH → ZERO_COST (0.0) → niet gevonden (-1.0)
    """
    # Rift-only / unobtainables → prijs 0, recept blijft geldig
    if ing_id in ZERO_COST_ITEMS or ing_id.upper() in ZERO_COST_ITEMS:
        return 0.0, _clean_name(ing_id)

    # Bazaar — probeer meerdere varianten
    for candidate in [
        ing_id,
        ing_id.upper(),
        _normalize_ingredient_id(ing_id),
        _normalize_ingredient_id(ing_id.upper()),
    ]:
        if candidate in bazaar:
            return bazaar[candidate].buy_price, bazaar[candidate].display_name

    # NPC prijs
    npc = NPC_PRICES.get(ing_id) or NPC_PRICES.get(ing_id.upper())
    if npc:
        return float(npc), _clean_name(ing_id)

    # AH prijs — gebruik mediaan van actieve listings
    if ah_prices:
        name = _clean_name(ing_id)
        ah_data = ah_prices.get(name) or ah_prices.get(_clean_name(ing_id.upper()))
        if ah_data:
            price = ah_data["price"] if isinstance(ah_data, dict) else float(ah_data)
            return price, name

    return -1.0, _clean_name(ing_id)  # niet gevonden


def _fmt(n: float) -> str:
    if n >= 1_000_000_000: return f"{n/1_000_000_000:.2f}B"
    if n >= 1_000_000:     return f"{n/1_000_000:.2f}M"
    if n >= 1_000:         return f"{n/1_000:.1f}K"
    return f"{int(n):,}"


def _timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d %H-%M-%S")


EXPORT_DIR = Path("Exports")

_EXT_FOLDER = {
    "csv":  "CSV",
    "json": "JSON",
    "xlsx": "Excel",
    "md":   "Markdown",
}

_MODE_FOLDER = {
    "craft": "Bazaar Craft Flipper",
    "flip":  "Bazaar Flipper",
    "ah":    "AH Craft Flipper",
}

def _export_filename(mode: str, sort_col: str, ext: str) -> Path:
    subfolder = EXPORT_DIR / _EXT_FOLDER.get(ext, ext.upper()) / _MODE_FOLDER.get(mode, mode)
    subfolder.mkdir(parents=True, exist_ok=True)
    return subfolder / f"Hypixel {_MODE_FOLDER.get(mode, mode)} - {sort_col} - {_timestamp()}.{ext}"


def _log_norm(values: list) -> list:
    log_vals = [math.log10(max(v, 1)) for v in values]
    mn, mx = min(log_vals), max(log_vals)
    r = mx - mn or 1
    return [(v - mn) / r for v in log_vals]


def _compute_scores(items: list, profit_fn, volume_fn, margin_fn) -> None:
    if not items:
        return
    lp = _log_norm([profit_fn(x) for x in items])
    lv = _log_norm([volume_fn(x) for x in items])
    lm = _log_norm([margin_fn(x) for x in items])
    for i, item in enumerate(items):
        item.score = WEIGHT_PROFIT * lp[i] + WEIGHT_VOLUME * lv[i] + WEIGHT_MARGIN * lm[i]


# ══════════════════════════════════════════════════════
#  HYPIXEL API
# ══════════════════════════════════════════════════════

class HypixelAPI:
    BASE = "https://api.hypixel.net"

    def __init__(self, api_key: str):
        self.key     = api_key
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": "HypixelBazaarTool/1.0"})

    def _get(self, endpoint: str, params: dict = None) -> dict:
        p = {"key": self.key}
        if params:
            p.update(params)
        resp = self.session.get(f"{self.BASE}/{endpoint}", params=p, timeout=15)
        resp.raise_for_status()
        data = resp.json()
        if not data.get("success", True):
            raise RuntimeError(f"API fout: {data.get('cause', 'onbekend')}")
        return data

    def get_bazaar(self) -> dict[str, BazaarItem]:
        """Haal alle bazaar items op."""
        data = self._get("v2/skyblock/bazaar")
        result = {}
        for item_id, product in data.get("products", {}).items():
            qs = product.get("quick_status", {})
            # Sla items over zonder prijsdata
            if not qs.get("buyPrice") and not qs.get("sellPrice"):
                continue
            result[item_id] = BazaarItem(
                item_id      = item_id,
                display_name = _clean_name(item_id),
                buy_price    = qs.get("buyPrice", 0),
                sell_price   = qs.get("sellPrice", 0),
                buy_volume   = qs.get("buyMovingWeek", 0),
                sell_volume  = qs.get("sellMovingWeek", 0),
            )
        return result

    def get_ah_prices(self) -> dict[str, dict]:
        """
        Combineert twee endpoints:
        - /v2/skyblock/auctions        → actuele BIN prijzen (listing prices)
        - /v2/skyblock/auctions_ended  → recente verkopen (echte verkoopprijzen + sales count)

        Geeft dict terug: {item_name: {"price": float, "sales": int}}
        Gecached voor AH_CACHE_MINUTES minuten.
        """
        if AH_CACHE_FILE.exists():
            try:
                cached = json.loads(AH_CACHE_FILE.read_text(encoding="utf-8"))
                age = datetime.now() - datetime.fromisoformat(cached["timestamp"])
                if age < timedelta(minutes=AH_CACHE_MINUTES):
                    print(f"  AH cache gebruikt ({int(age.total_seconds()//60)}min oud)")
                    return cached["prices"]
            except Exception:
                pass

        # ── Actieve BIN veilingen ophalen ──────────────────────────────────
        # auctions_ended heeft geen item_name (alleen NBT bytes), dus we
        # gebruiken het aantal actieve listings als proxy voor populariteit.
        print("  Actieve BIN veilingen ophalen (pagina 0)...")
        bin_prices  = defaultdict(list)   # name → [prijzen]
        bin_count   = defaultdict(int)    # name → aantal actieve listings

        try:
            data = self._get("v2/skyblock/auctions", {"page": 0})
            total_pages = data.get("totalPages", 1)
            print(f"  {total_pages} pagina's — even geduld...")

            def process_page(auctions):
                for auction in auctions:
                    if not auction.get("bin", False):
                        continue
                    name = _strip_color(auction.get("item_name", "")).strip()
                    price = auction.get("starting_bid") or auction.get("price", 0)
                    if name and price > 0:
                        bin_prices[name].append(price)
                        bin_count[name] += 1

            process_page(data.get("auctions", []))

            for page in range(1, min(total_pages, 50)):
                if page % 10 == 0:
                    print(f"  Pagina {page}/{min(total_pages, 50)}...")
                try:
                    pdata = self._get("v2/skyblock/auctions", {"page": page})
                    process_page(pdata.get("auctions", []))
                    time.sleep(0.05)
                except Exception as pe:
                    print(f"  Pagina {page} overgeslagen: {pe}")
                    break
        except Exception as e:
            print(f"  Actieve veilingen mislukt: {e}")

        # ── Mediaan prijs + listing count per item ──────────────────────────
        prices = {}
        for name, vals in bin_prices.items():
            sorted_vals = sorted(vals)
            mid = len(sorted_vals) // 2
            prices[name] = {
                "price":    sorted_vals[mid],
                "sales":    bin_count[name],   # actieve listings = populariteitsproxy
            }

        AH_CACHE_FILE.write_text(json.dumps({
            "timestamp": datetime.now().isoformat(),
            "prices": prices,
        }), encoding="utf-8")
        print(f"  {len(prices)} unieke AH items verwerkt")
        return prices


def _clean_name(item_id: str) -> str:
    """Zet ENCHANTED_WHEAT om naar 'Enchanted Wheat'."""
    return item_id.replace("_", " ").title()


def _strip_color(text: str) -> str:
    """Verwijder Minecraft kleurcodes (§a, §l etc.)."""
    return re.sub(r"§.", "", text)


# ══════════════════════════════════════════════════════
#  RECEPTEN LADEN (NEU REPO)
# ══════════════════════════════════════════════════════


def load_recipes(api_key: str = "") -> dict[str, dict]:
    """
    Laad craft recepten uit de NEU GitHub repo.
    NEU slot formaat: "count:ITEM_ID" of "ITEM_ID:damage"
      - "32:ENCHANTED_DIAMOND" → 32x Enchanted Diamond
      - "ENCHANTED_DIAMOND:0"  → 1x Enchanted Diamond (0 = damage value)
    """
    if RECIPE_CACHE_FILE.exists():
        try:
            cached = json.loads(RECIPE_CACHE_FILE.read_text(encoding="utf-8"))
            age = datetime.now() - datetime.fromisoformat(cached["timestamp"])
            if age < timedelta(hours=RECIPE_CACHE_HOURS):
                print(f"  Recepten cache gebruikt ({int(age.total_seconds()//3600)}u oud, {len(cached['recipes'])} items)")
                # Laad ook gecachede NPC prijzen
                if "npc_prices" in cached:
                    NPC_PRICES.update(cached["npc_prices"])
                return cached["recipes"]
        except Exception:
            pass

    print("  NEU recepten downloaden (eenmalig, ~18MB)...")
    resp = requests.get(NEU_ZIP_URL, stream=True, timeout=120)
    resp.raise_for_status()

    total = int(resp.headers.get("content-length", 0))
    downloaded = 0
    chunks = []
    for chunk in resp.iter_content(65536):
        chunks.append(chunk)
        downloaded += len(chunk)
        if total:
            pct = downloaded / total * 100
            print(f"\r  Downloaden: {pct:.0f}%  ({downloaded//1024//1024}MB/{total//1024//1024}MB)", end="", flush=True)
    print()

    z = zipfile.ZipFile(io.BytesIO(b"".join(chunks)))
    item_files = [f for f in z.namelist()
                  if "/items/" in f and f.endswith(".json") and not f.endswith("/")]

    print(f"  {len(item_files)} item bestanden parsen...")
    recipes = {}
    SLOTS = ["A1","A2","A3","B1","B2","B3","C1","C2","C3"]

    for filepath in item_files:
        try:
            data = json.loads(z.read(filepath).decode("utf-8"))
        except Exception:
            continue

        item_id = data.get("internalname", "")
        if not item_id:
            continue

        recipe = data.get("recipe")
        if not recipe:
            continue

        ingredient_counts = defaultdict(int)
        for slot in SLOTS:
            val = recipe.get(slot, "")
            if not val:
                continue

            parts = val.split(":")
            if len(parts) == 2:
                ing_id = parts[0].strip()
                # "ENCHANTED_DIAMOND:32" → count=32
                # "ENCHANTED_DIAMOND:0"  → count=1 (0 betekent gewoon 1x)
                raw_count = int(parts[1]) if parts[1].isdigit() else 0
                count = raw_count if raw_count > 0 else 1
            else:
                ing_id = val.strip()
                count  = 1

            if ing_id:
                ingredient_counts[ing_id] += count

        if not ingredient_counts:
            continue

        output_count = int(recipe.get("count", 1))
        display = _strip_color(data.get("displayname", _clean_name(item_id)))

        recipes[item_id] = {
            "ingredients":  list(ingredient_counts.items()),
            "output_count": output_count,
            "display_name": display,
        }

    # Haal ook NPC prijzen op via Hypixel Items API
    print("  NPC prijzen ophalen...")
    npc_prices_live = {}
    try:
        params = {}
        if api_key:
            params["key"] = api_key
        resp2 = requests.get("https://api.hypixel.net/v2/resources/skyblock/items",
                             params=params, timeout=30)
        if resp2.ok:
            for item in resp2.json().get("items", []):
                npc_price = item.get("npc_sell_price")
                if npc_price and npc_price > 0:
                    npc_prices_live[item["id"]] = npc_price
            print(f"  {len(npc_prices_live)} NPC prijzen opgehaald")
    except Exception as e:
        print(f"  NPC prijzen ophalen mislukt: {e} (gebruik hardcoded fallback)")

    # Merge live NPC prijzen in de globale NPC_PRICES dict
    NPC_PRICES.update(npc_prices_live)

    RECIPE_CACHE_FILE.write_text(json.dumps({
        "timestamp": datetime.now().isoformat(),
        "recipes":   recipes,
        "npc_prices": NPC_PRICES,
    }, ensure_ascii=False), encoding="utf-8")
    print(f"  {len(recipes)} recepten gecached in {RECIPE_CACHE_FILE}")
    return recipes


# ══════════════════════════════════════════════════════
#  ANALYSE: CRAFT FLIPS
# ══════════════════════════════════════════════════════

def analyze_craft_flips(bazaar: dict, recipes: dict,
                        ah_prices: dict = None) -> list[CraftFlip]:
    """
    Koop materials via buy order → craft → verkoop via sell offer.
    Alleen items die zowel craftbaar zijn als in de bazaar staan.
    """
    results = []
    _dbg_missing_ids = defaultdict(int)

    for item_id, recipe in recipes.items():
        # Gecraftte item moet in bazaar staan
        if item_id not in bazaar:
            continue

        crafted = bazaar[item_id]
        output_count = recipe["output_count"]

        # Bereken kosten per ingredient
        total_cost = 0.0
        ingredient_details = []
        valid = True

        for ing_id, amount_per_craft in recipe["ingredients"]:
            price, ing_name = _get_ingredient_price(ing_id, bazaar, ah_prices)
            if price < 0:
                _dbg_missing_ids[ing_id] += 1
                valid = False
                break

            cost_this = price * amount_per_craft
            total_cost += cost_this
            ingredient_details.append((ing_name, amount_per_craft, price))

        if not valid or total_cost <= 0:
            continue

        # Kosten per gecraft item
        cost_per_item = total_cost / output_count

        # Opbrengst: we plaatsen een sell offer op buy_price niveau (insta-buy prijs)
        # Dit is de prijs die kopers betalen — hoger dan sell_price (buy order prijs)
        output_value = crafted.buy_price * (1 - BAZAAR_SELL_TAX)

        profit = output_value - cost_per_item
        if profit <= 0:
            continue

        margin_pct = profit / cost_per_item * 100
        volume = crafted.sell_volume  # orders/week van het gecraftte item

        results.append(CraftFlip(
            name         = recipe["display_name"],
            item_id      = item_id,
            input_cost   = cost_per_item,
            output_value = output_value,
            profit       = profit,
            margin_pct   = margin_pct,
            volume       = volume,
            ingredients  = ingredient_details,
        ))

    in_bz = sum(1 for i in recipes if i in bazaar)
    top_missing = sorted(_dbg_missing_ids.items(), key=lambda x: x[1], reverse=True)[:10]
    print(f"  Craft: {in_bz} in bazaar | {len(results)} winstgevend | "
          f"{sum(_dbg_missing_ids.values())} ingredient mismatches"
          + (f" | ontbrekend: {[x[0] for x in top_missing]}" if top_missing else ""))


    _compute_scores(results,
                    profit_fn=lambda x: x.profit,
                    volume_fn=lambda x: x.volume,
                    margin_fn=lambda x: x.margin_pct)

    return sorted(results, key=lambda x: x.score, reverse=True)


# ══════════════════════════════════════════════════════
#  ANALYSE: BAZAAR FLIPS
# ══════════════════════════════════════════════════════

def analyze_bazaar_flips(bazaar: dict) -> list[BazaarFlip]:
    """
    Koop via buy order (sell_price niveau) → wacht op vulling →
    verkoop via sell offer (buy_price niveau) → na tax = profit.
    """
    results = []

    for item_id, item in bazaar.items():
        if item.buy_price <= 0 or item.sell_price <= 0:
            continue
        if item.buy_price <= item.sell_price:
            continue

        # Profit per item na tax
        profit = item.sell_price * (1 - BAZAAR_SELL_TAX) - item.buy_price
        # Wacht — we kopen via buy order (betalen sell_price niveau)
        # en verkopen via sell offer (ontvangen buy_price niveau minus tax)
        # Eigenlijk: profit = buy_price*(1-tax) - sell_price
        profit = item.buy_price * (1 - BAZAAR_SELL_TAX) - item.sell_price

        if profit <= 0:
            continue

        margin_pct = item.margin_pct()
        volume = min(item.buy_volume, item.sell_volume)

        results.append(BazaarFlip(
            item_id     = item_id,
            name        = item.display_name,
            buy_order   = item.sell_price,
            sell_offer  = item.buy_price,
            profit      = profit,
            margin_pct  = margin_pct,
            volume      = volume,
            buy_volume  = item.buy_volume,
            sell_volume = item.sell_volume,
        ))

    _compute_scores(results,
                    profit_fn=lambda x: x.profit,
                    volume_fn=lambda x: x.volume,
                    margin_fn=lambda x: x.margin_pct)

    return sorted(results, key=lambda x: x.score, reverse=True)


# ══════════════════════════════════════════════════════
#  ANALYSE: BAZAAR → CRAFT → AH
# ══════════════════════════════════════════════════════

def analyze_ah_craft_flips(bazaar: dict, recipes: dict,
                            ah_prices: dict) -> list[AHCraftFlip]:
    """
    Koop materials van bazaar (insta-buy) → craft → verkoop op AH.
    """
    # Bouw een mapping van display_name → ah_price
    # AH gebruikt display names, recepten gebruiken item IDs
    results = []

    for item_id, recipe in recipes.items():
        # Item mag NIET in bazaar staan (anders gebruik je craft flip)
        # OF mag wel in bazaar staan maar AH is winstgevender
        display = recipe["display_name"]

        # Zoek AH prijs (op display name)
        ah_data = ah_prices.get(display) or ah_prices.get(_clean_name(item_id))
        if not ah_data:
            continue
        # Ondersteun zowel nieuw dict formaat als oud float formaat
        if isinstance(ah_data, dict):
            ah_price = ah_data["price"]
            ah_sales_count = ah_data["sales"]
        else:
            ah_price = float(ah_data)
            ah_sales_count = 1

        output_count = recipe["output_count"]
        total_cost = 0.0
        ingredient_details = []
        valid = True

        for ing_id, amount_per_craft in recipe["ingredients"]:
            price, ing_name = _get_ingredient_price(ing_id, bazaar, ah_prices)
            if price < 0:
                valid = False
                break

            cost_this = price * amount_per_craft
            total_cost += cost_this
            ingredient_details.append((ing_name, amount_per_craft, price))

        if not valid or total_cost <= 0:
            continue

        cost_per_item = total_cost / output_count
        net_ah = ah_price * (1 - AH_TAX)
        profit = net_ah - cost_per_item

        if profit <= 0:
            continue

        margin_pct = profit / cost_per_item * 100

        ah_sales = ah_sales_count

        results.append(AHCraftFlip(
            name          = display,
            item_id       = item_id,
            material_cost = cost_per_item,
            ah_price      = ah_price,
            profit        = profit,
            margin_pct    = margin_pct,
            ah_sales      = ah_sales,
            ingredients   = ingredient_details,
        ))

    _compute_scores(results,
                    profit_fn=lambda x: x.profit,
                    volume_fn=lambda x: x.ah_sales,
                    margin_fn=lambda x: x.margin_pct)

    return sorted(results, key=lambda x: x.score, reverse=True)


# ══════════════════════════════════════════════════════
#  EXPORTS
# ══════════════════════════════════════════════════════

def _rows_craft(items: list) -> tuple:
    headers = ["#", "Naam", "Cost/item", "Output", "Profit", "Margin", "Volume/wk", "Score", "Ingredients"]
    rows = []
    for i, c in enumerate(items, 1):
        ings = " | ".join(f"{a}x {n}" for n, a, _ in c.ingredients)
        rows.append([i, c.name, c.cost_str(), c.output_str(), c.profit_str(),
                     c.margin_str(), c.volume_str(), c.score_str(), ings])
    return headers, rows


def _rows_flip(items: list) -> tuple:
    headers = ["#", "Naam", "Buy order", "Sell offer", "Profit/item",
               "Margin", "Buy vol/wk", "Sell vol/wk", "Min vol/wk", "Score"]
    rows = []
    for i, f in enumerate(items, 1):
        rows.append([i, f.name, f.buy_str(), f.sell_str(), f.profit_str(),
                     f.margin_str(), f.buy_vol_str(), f.sell_vol_str(),
                     f.volume_str(), f.score_str()])
    return headers, rows


def _rows_ah(items: list) -> tuple:
    headers = ["#", "Naam", "Material cost", "AH prijs", "Profit", "Margin", "AH sales", "Score", "Ingredients"]
    rows = []
    for i, a in enumerate(items, 1):
        ings = " | ".join(f"{amt}x {n}" for n, amt, _ in a.ingredients)
        rows.append([i, a.name, a.cost_str(), a.ah_str(), a.profit_str(),
                     a.margin_str(), str(a.ah_sales), a.score_str(), ings])
    return headers, rows


def export_csv(items: list, mode: str, sort_col: str, row_fn) -> Path:
    path = _export_filename(mode, sort_col, "csv")
    headers, rows = row_fn(items)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)
    return path


def export_json_file(items: list, mode: str, sort_col: str) -> Path:
    path = _export_filename(mode, sort_col, "json")

    def serialize(item):
        d = item.__dict__.copy()
        if "ingredients" in d:
            d["ingredients"] = [{"name": n, "amount": a, "price": p}
                                 for n, a, p in d["ingredients"]]
        return d

    path.write_text(json.dumps([serialize(x) for x in items],
                                indent=2, ensure_ascii=False), encoding="utf-8")
    return path


def export_excel(items: list, mode: str, sort_col: str, row_fn) -> Path:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    path = _export_filename(mode, sort_col, "xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = mode

    hdr_fill  = PatternFill("solid", fgColor="1E3A5F")
    hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    even_fill = PatternFill("solid", fgColor="E8F5E9")
    odd_fill  = PatternFill("solid", fgColor="F5F5F5")
    top_fill  = PatternFill("solid", fgColor="1B3A1B")
    b         = Side(style="thin", color="CCCCCC")
    border    = Border(left=b, right=b, top=b, bottom=b)

    headers, rows = row_fn(items)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = max(12, len(h) + 2)

    for ri, row in enumerate(rows, 2):
        fill = top_fill if ri <= 11 else (even_fill if ri % 2 == 0 else odd_fill)
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    wb.save(path)
    return path

def export_markdown(items: list, mode: str, sort_col: str, row_fn) -> Path:
    path = _export_filename(mode, sort_col, "md")
    headers, rows = row_fn(items)
    lines = [
        f"# Hypixel SkyBlock – {mode}",
        f"_Geëxporteerd op {datetime.now().strftime('%d-%m-%Y %H:%M')} · gesorteerd op: {sort_col}_\n",
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(str(v).replace("|", "\\|") for v in row) + " |")
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


# ══════════════════════════════════════════════════════
#  GUI
# ══════════════════════════════════════════════════════

DARK_BG    = "#1a1a2e"
PANEL_BG   = "#16213e"
ACCENT     = "#0f3460"
GOLD       = "#e2b96f"
GREEN      = "#4caf50"
PURPLE     = "#533483"
TEXT       = "#e0e0e0"
SUBTEXT    = "#aaaaaa"


class HypixelBazaarGUI:

    def __init__(self):
        self.api      = HypixelAPI(HYPIXEL_API_KEY)
        self.bazaar   = {}
        self.recipes  = {}
        self.ah       = {}

        self.craft_flips = []
        self.bz_flips    = []
        self.ah_flips    = []

        self.root = tk.Tk()
        self.root.title("Hypixel Bazaar Tool")
        self.root.geometry("1500x820")
        self.root.configure(bg=DARK_BG)

        self._setup_styles()
        self._build_ui()
        self.root.after(100, self._initial_load)

    def _setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")

        style.configure("TNotebook",           background=DARK_BG, borderwidth=0)
        style.configure("TNotebook.Tab",       background=ACCENT,  foreground=TEXT,
                         font=("Segoe UI", 10, "bold"), padding=[14, 6])
        style.map("TNotebook.Tab",
                  background=[("selected", GOLD)],
                  foreground=[("selected", "#1a1a2e")])

        style.configure("Treeview",
                         background=DARK_BG, foreground=TEXT,
                         fieldbackground=DARK_BG, rowheight=26,
                         font=("Segoe UI", 10))
        style.configure("Treeview.Heading",
                         background=PANEL_BG, foreground=GOLD,
                         font=("Segoe UI", 10, "bold"), relief="flat")
        style.map("Treeview", background=[("selected", ACCENT)])

    def _build_ui(self):
        # ── Header ──────────────────────────────────
        hdr = tk.Frame(self.root, bg=PANEL_BG, pady=10)
        hdr.pack(fill="x")

        tk.Label(hdr, text="🪙  Hypixel SkyBlock Bazaar Tool",
                 font=("Segoe UI", 17, "bold"), fg=GOLD, bg=PANEL_BG).pack(side="left", padx=16)

        self.status_var = tk.StringVar(value="Laden...")
        tk.Label(hdr, textvariable=self.status_var,
                 font=("Segoe UI", 10), fg=SUBTEXT, bg=PANEL_BG).pack(side="right", padx=16)

        tk.Button(hdr, text="↻  AH vernieuwen",
                  command=self._refresh_force_ah,
                  bg="#e67e22", fg="white",
                  font=("Segoe UI", 10, "bold"),
                  relief="flat", padx=10).pack(side="right", padx=4)

        self.refresh_btn = tk.Button(hdr, text="↻  Bazaar vernieuwen",
                                      command=self._refresh,
                                      bg=GREEN, fg="white",
                                      font=("Segoe UI", 10, "bold"),
                                      relief="flat", padx=12)
        self.refresh_btn.pack(side="right", padx=8)

        # ── Notebook ─────────────────────────────────
        self.nb = ttk.Notebook(self.root)
        self.nb.pack(fill="both", expand=True, padx=8, pady=8)

        self.tab_craft = self._make_tab("⚒  Craft Flips",    self._craft_columns())
        self.tab_flip  = self._make_tab("📈  Bazaar Flips",  self._flip_columns())
        self.tab_ah    = self._make_tab("🏷  AH Craft Flips", self._ah_columns())

        # ── Filter bar per tab ───────────────────────
        self._add_filter_bar(self.tab_craft, "craft",
            [("Min Profit", MIN_CRAFT_PROFIT), ("Min Volume/wk", MIN_CRAFT_VOLUME)])
        self._add_filter_bar(self.tab_flip, "flip",
            [("Min Profit", MIN_FLIP_PROFIT), ("Min Margin %", MIN_FLIP_MARGIN),
             ("Min Volume/wk", MIN_FLIP_VOLUME)])
        self._add_filter_bar(self.tab_ah, "ah",
            [("Min Profit", MIN_AH_PROFIT)])

    def _make_tab(self, label: str, columns: list) -> tk.Frame:
        frame = tk.Frame(self.nb, bg=DARK_BG)
        self.nb.add(frame, text=label)
        return frame

    def _craft_columns(self):
        return [("rank","#",50),("name","Naam",260),("cost","Cost/item",100),
                ("output","Output",100),("profit","Profit",110),("margin","Margin",80),
                ("volume","Volume/wk",100),("score","Score",70),("ings","Ingredients",320)]

    def _flip_columns(self):
        return [("rank","#",50),("name","Naam",220),("buy","Buy order",100),
                ("sell","Sell offer",100),("profit","Profit/item",100),
                ("margin","Margin",75),("buyvol","Buy vol/wk",100),
                ("sellvol","Sell vol/wk",100),("volume","Min vol/wk",100),
                ("score","Score",70)]

    def _ah_columns(self):
        return [("rank","#",50),("name","Naam",260),("cost","Material cost",120),
                ("ah","AH prijs",110),("profit","Profit",110),("margin","Margin",80),
                ("sales","AH sales",90),("score","Score",70),("ings","Ingredients",300)]

    def _add_filter_bar(self, parent: tk.Frame, mode: str, filters: list):
        bar = tk.Frame(parent, bg=ACCENT, pady=6)
        bar.pack(fill="x", padx=0, pady=(0,4))

        vars_ = {}
        for label, default in filters:
            tk.Label(bar, text=label+":", fg=TEXT, bg=ACCENT,
                     font=("Segoe UI", 10)).pack(side="left", padx=(12,4))
            v = tk.StringVar(value=str(default))
            vars_[label] = v
            tk.Entry(bar, textvariable=v, width=10,
                     font=("Segoe UI", 10), bg=DARK_BG, fg=TEXT,
                     insertbackground=TEXT, relief="flat").pack(side="left", padx=(0,10))

        tk.Button(bar, text="Filter toepassen",
                  command=lambda m=mode, vs=vars_: self._apply_filter(m, vs),
                  bg=GOLD, fg=DARK_BG, font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=10).pack(side="left")

        tk.Button(bar, text="✕  Filters wissen",
                  command=lambda m=mode, vs=vars_: self._clear_filters(m, vs),
                  bg="#c0392b", fg="white", font=("Segoe UI", 9, "bold"),
                  relief="flat", padx=10).pack(side="left", padx=(6, 0))

        # Export knoppen
        exp = tk.Frame(bar, bg=ACCENT)
        exp.pack(side="right", padx=12)
        for lbl, ext in [("CSV","csv"),("JSON","json"),("Excel","xlsx"),("Markdown","md")]:
            tk.Button(exp, text=lbl,
                      command=lambda m=mode, e=ext: self._export(m, e),
                      bg=PURPLE, fg="white", font=("Segoe UI", 9, "bold"),
                      relief="flat", padx=8, pady=2).pack(side="left", padx=3)

        # Treeview
        tree_frame = tk.Frame(parent, bg=DARK_BG)
        tree_frame.pack(fill="both", expand=True, padx=4, pady=4)

        cols = {"craft": self._craft_columns(),
                "flip":  self._flip_columns(),
                "ah":    self._ah_columns()}[mode]

        tree = ttk.Treeview(tree_frame, columns=[c[0] for c in cols],
                             show="headings", selectmode="browse")

        for col_id, col_label, col_width in cols:
            tree.heading(col_id, text=col_label,
                         command=lambda c=col_id, m=mode: self._sort_tree(m, c))
            tree.column(col_id, width=col_width, minwidth=40,
                        anchor="e" if col_id not in ("name","ings","rank") else "w")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",   command=tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        tree.tag_configure("top",  background="#1b3a1b")
        tree.tag_configure("even", background=DARK_BG)
        tree.tag_configure("odd",  background=PANEL_BG)
        tree.bind("<Double-1>", lambda e, m=mode: self._show_detail(m, e))

        setattr(self, f"tree_{mode}", tree)
        setattr(self, f"filter_vars_{mode}", vars_)

    def _load_tree(self, mode: str, items: list, row_fn):
        tree = getattr(self, f"tree_{mode}")
        tree.delete(*tree.get_children())
        _, rows = row_fn(items)
        for i, row in enumerate(rows):
            tag = "top" if i < 10 else ("even" if i % 2 == 0 else "odd")
            tree.insert("", "end", tags=(tag,), values=row)

    def _clear_filters(self, mode: str, vars_: dict):
        """Zet alle filterwaarden op 0 en laad volledige lijst."""
        for v in vars_.values():
            v.set("0")
        self._apply_filter(mode, vars_)

    def _apply_filter(self, mode: str, vars_: dict):
        def get_float(label):
            try: return float(vars_[label].get() or 0)
            except: return 0

        if mode == "craft":
            min_p = get_float("Min Profit")
            min_v = get_float("Min Volume/wk")
            filtered = [c for c in self.craft_flips if c.profit >= min_p and c.volume >= min_v]
            self._load_tree("craft", filtered, _rows_craft)
            self.status_var.set(f"Craft Flips: {len(filtered)} resultaten")

        elif mode == "flip":
            min_p = get_float("Min Profit")
            min_m = get_float("Min Margin %")
            min_v = get_float("Min Volume/wk")
            filtered = [f for f in self.bz_flips
                        if f.profit >= min_p and f.margin_pct >= min_m and f.volume >= min_v]
            self._load_tree("flip", filtered, _rows_flip)
            self.status_var.set(f"Bazaar Flips: {len(filtered)} resultaten")

        elif mode == "ah":
            min_p = get_float("Min Profit")
            filtered = [a for a in self.ah_flips if a.profit >= min_p]
            self._load_tree("ah", filtered, _rows_ah)
            self.status_var.set(f"AH Craft Flips: {len(filtered)} resultaten")

    def _sort_tree(self, mode: str, col_id: str):
        items = {"craft": self.craft_flips, "flip": self.bz_flips, "ah": self.ah_flips}[mode]
        row_fn = {"craft": _rows_craft, "flip": _rows_flip, "ah": _rows_ah}[mode]

        sort_map = {
            "rank": "score", "name": "score", "score": "score",
            "profit": "profit", "margin": "margin_pct", "volume": "volume",
            "cost": "input_cost", "output": "output_value",
            "buy": "buy_order", "sell": "sell_offer",
            "ah": "ah_price", "sales": "ah_sales",
        }
        attr = sort_map.get(col_id, "score")
        sorted_items = sorted(items, key=lambda x: getattr(x, attr, 0), reverse=True)
        self._load_tree(mode, sorted_items, row_fn)

    def _show_detail(self, mode: str, event):
        tree = getattr(self, f"tree_{mode}")
        item = tree.focus()
        if not item:
            return
        vals = tree.item(item, "values")
        if not vals:
            return

        if mode in ("craft", "ah"):
            # Last column is ingredients
            name = vals[1]
            ings = vals[-1]
            info = f"{name}\n{'─'*50}\n"
            for part in vals[2:-1]:
                info += f"{part}\n"
            info += f"\nIngredients:\n"
            for ing in ings.split(" | "):
                info += f"  • {ing}\n"
        else:
            name = vals[1]
            info = f"{name}\n{'─'*50}\n"
            for v in vals[2:]:
                info += f"{v}\n"

        messagebox.showinfo(name, info)

    def _export(self, mode: str, ext: str):
        items = {"craft": self.craft_flips, "flip": self.bz_flips, "ah": self.ah_flips}[mode]
        row_fn = {"craft": _rows_craft, "flip": _rows_flip, "ah": _rows_ah}[mode]

        if ext == "csv":
            path = export_csv(items, mode, "score", row_fn)
        elif ext == "json":
            path = export_json_file(items, mode, "score")
        elif ext == "xlsx":
            path = export_excel(items, mode, "score", row_fn)
        elif ext == "md":
            path = export_markdown(items, mode, "score", row_fn)
        else:
            return

        if path:
            messagebox.showinfo("Export geslaagd", f"Opgeslagen als:\n{path.resolve()}")

    def _initial_load(self):
        self.status_var.set("Data ophalen...")
        self.refresh_btn.config(state="disabled")
        threading.Thread(target=self._load_data, daemon=True).start()

    def _refresh(self, force_ah=False):
        """
        Vernieuw bazaar + analyse. AH cache wordt alleen gewist als force_ah=True
        of als de cache ouder is dan AH_CACHE_MINUTES (automatisch in get_ah_prices).
        """
        self.refresh_btn.config(state="disabled")
        self.status_var.set("Vernieuwen...")
        if force_ah and AH_CACHE_FILE.exists():
            AH_CACHE_FILE.unlink()
            self.status_var.set("Vernieuwen (AH cache gewist)...")
        threading.Thread(target=self._load_data, daemon=True).start()

    def _refresh_force_ah(self):
        self._refresh(force_ah=True)

    def _load_data(self):
        try:
            self.root.after(0, lambda: self.status_var.set("Recepten laden..."))
            self.recipes = load_recipes(HYPIXEL_API_KEY)

            self.root.after(0, lambda: self.status_var.set("Bazaar data ophalen..."))
            self.bazaar = self.api.get_bazaar()

            self.root.after(0, lambda: self.status_var.set("AH prijzen ophalen..."))
            self.ah = self.api.get_ah_prices()

            self.root.after(0, lambda: self.status_var.set("Analyseren..."))

            self.craft_flips = analyze_craft_flips(self.bazaar, self.recipes, self.ah)
            self.bz_flips    = analyze_bazaar_flips(self.bazaar)
            self.ah_flips    = analyze_ah_craft_flips(self.bazaar, self.recipes, self.ah)

            self.root.after(0, self._update_ui)

        except Exception as e:
            err = str(e)
            self.root.after(0, lambda err=err: messagebox.showerror("Fout", err))
            self.root.after(0, lambda err=err: self.status_var.set(f"Fout: {err}"))
        finally:
            self.root.after(0, lambda: self.refresh_btn.config(state="normal"))

    def _update_ui(self):
        self._load_tree("craft", self.craft_flips, _rows_craft)
        self._load_tree("flip",  self.bz_flips,    _rows_flip)
        self._load_tree("ah",    self.ah_flips,     _rows_ah)

        now = datetime.now().strftime("%H:%M:%S")
        self.status_var.set(
            f"✓ Laatste update: {now}  |  "
            f"Craft: {len(self.craft_flips)}  |  "
            f"Flips: {len(self.bz_flips)}  |  "
            f"AH: {len(self.ah_flips)}"
        )
        print(f"\n✓ Analyse klaar — "
              f"{len(self.craft_flips)} craft flips, "
              f"{len(self.bz_flips)} bazaar flips, "
              f"{len(self.ah_flips)} AH craft flips")

    def run(self):
        self.root.mainloop()


# ══════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════

if __name__ == "__main__":
    print()
    print("  ╔════════════════════════════════════════════╗")
    print("  ║   Hypixel SkyBlock Bazaar Tool  🪙        ║")
    print("  ╚════════════════════════════════════════════╝")
    print()

    if not HYPIXEL_API_KEY:
        print("  Geen API key gevonden!")
        print("  Voeg toe aan .env:  HYPIXEL_API_KEY=jouw_key_hier")
        print("  Key krijgen: join Hypixel → /api new in-game")
        exit(1)

    print(f"  API key geladen ({'*'*8}{HYPIXEL_API_KEY[-4:]})")
    print()

    gui = HypixelBazaarGUI()
    gui.run()